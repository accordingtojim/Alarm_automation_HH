from openpyxl import load_workbook
from openpyxl import Workbook
import shutil
import os
import config
def file_creation_2( path_to_template_AUX):
    counter = 0
    local_type_HH = config.convert_to_int(config.array_type_HH)
    local_HH_GUI = config.convert_to_int(config.num_HH_GUI)
    local_n_PI = config.convert_to_int(config.n_PI)
    path_to_new_template = (path_to_template_AUX.replace('.xlsx','')) + '_new_AUX' + '.xlsx'
    config.global_list.append(path_to_new_template)
    shutil.copy (path_to_template_AUX , path_to_new_template)
    wb = load_workbook(path_to_new_template)    
    ws = wb.active
    number_row = 1
    number_column = 1
    while not ((ws.cell(row = 1 ,column = number_column).value == None) and (ws.cell(row = 1,column = number_column + 1).value == None)) :
        number_column += 1
    number_column -= 1
    while not ((ws.cell(row=number_row,column=1).value == None) and (ws.cell(row=number_row + 1,column=1).value == None)) :
        number_row += 1
    number_row -= 1
    array_parsed3 = config.p_array_arrangement(local_type_HH,local_HH_GUI)
    for n in range(1,local_HH_GUI*2):
            for j in range(1,number_column+1):
                for i in range(1,number_row+1):
                    ws.cell(row=i+n*number_row,column=j).value=ws.cell(row=i,column=j).value
    #wb.save(path_to_new_template)
    for PI in range(1 , local_n_PI + 1):
        for HH in range(1 , array_parsed3[PI-1] + 1):
            if HH>9:
                name_HH = "HH1HD"
                name_PI = "PI"
            else:
                name_HH = "HH1HD0"
                name_PI = "PI0"
            for AUX in range(1 , 3):
                for i in range(1,number_row+1):
                    ws.cell(row=i+counter*number_row,column=1).value = str(ws.cell(row =i+counter*number_row, column = 1).value)\
                    + " | "\
                    + "AUX,PCS"\
                    + str(AUX)\
                    + " - "\
                    + name_HH\
                    + str(HH)\
                    + " - "\
                    + name_PI\
                    + str(HH)
                    ws.cell(row=i+counter*number_row,column=3).value = name_HH\
                    + str(HH)\
                    + "_PCS"\
                    + str(AUX)\
                    + "_AUX_Status_HMI."\
                    + str(ws.cell(row=i+counter*number_row,column=3).value)
                counter += 1
    ws.insert_cols(2)
    wb.save(path_to_new_template)