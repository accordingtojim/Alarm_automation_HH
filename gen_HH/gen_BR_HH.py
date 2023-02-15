from openpyxl import load_workbook
from openpyxl import Workbook
import shutil
import os
import config
def file_creation_6( path_to_template_BR): 
    counter = 0
    local_type_HH = config.convert_to_int(config.array_type_HH)
    local_n_BR = config.convert_to_int(config.n_battery_rack)
    local_n_BB = config.convert_to_int(config.n_battery_bank)
    local_HH_GUI = config.convert_to_int(config.num_HH_GUI)
    local_n_PI = config.convert_to_int(config.n_PI)
    path_to_new_template = (path_to_template_BR.replace('.xlsx','')) + '_new_BR' + '.xlsx'
    config.global_list.append(path_to_new_template)
    shutil.copy (path_to_template_BR , path_to_new_template)
    wb = load_workbook(path_to_new_template)    
    ws = wb.active
    number_row = 1
    number_column = 1
    while not ((ws.cell(row = 1 ,column = number_column).value == None) and (ws.cell(row = 1,column = number_column + 1).value == None)) :
        number_column += 1
    number_column -= 1
    while not ((ws.cell(row=number_row,column=1).value == None) and (ws.cell(row=number_row + 1,column=1).value == None)) :
        number_row += 1
    number_row -= 1
    array_parsed1 = config.p_array_arrangement(local_type_HH,local_n_BB)
    array_parsed2 = config.p_array_arrangement(local_type_HH,local_n_BR)
    array_parsed3 = config.p_array_arrangement(local_type_HH,local_HH_GUI)
    array_arranged3 = config.array_arrangement(array_parsed1,array_parsed2)
    
    for n in range(1,array_arranged3):
            for j in range(1,number_column+1):
                for i in range(1,number_row+1):
                    ws.cell(row=i+n*number_row,column=j).value=ws.cell(row=i,column=j).value
    #wb.save(path_to_new_template)
    for PI in range(1 , local_n_PI + 1):       
        for HH in range (1 , array_parsed3 [PI-1]+1): 
            if HH>9:
                name_HH = "HH1HD"
            else:
                name_HH = "HH1HD0"   
            for BB in range(1 , array_parsed1[PI-1]+1):
                for BR in range(1,array_parsed2[PI-1]+1):
                    if BR>9:
                        name_BR = "BR"
                    else:
                        name_BR = "BR0"
                    for i in range(1,number_row+1):
                        if ('Spare' or 'spare') in ws.cell(row=i+counter*number_row,column=1).value:
                            ws.cell(row=i+counter*number_row,column=1).value = str(ws.cell(row =i+counter*number_row, column = 1).value)\
                            + " | "\
                            + str(ws.cell(row =i+counter*number_row, column = 3).value) + "." + str(ws.cell(row =i+counter*number_row, column = 4).value)
                            ws.cell(row=i+counter*number_row,column=3).value = name_HH\
                            + str(HH)\
                            + "_BMS"\
                            + str(BB)\
                            + "_"\
                            + name_BR\
                            + str(BR)\
                            + "."\
                            + ws.cell(row=i+counter*number_row,column=3).value
                        else:
                            ws.cell(row=i+counter*number_row,column=1).value = str(ws.cell(row =i+counter*number_row, column = 1).value)\
                            + " | "\
                            + name_BR\
                            + str(BR)\
                            + ",BB0"\
                            + str(BB)\
                            + " - "\
                            + name_HH\
                            + str(HH)\
                            + " - PI0"\
                            + str(PI)
                            ws.cell(row=i+counter*number_row,column=3).value = name_HH\
                            + str(HH)\
                            + "_BMS"\
                            + str(BB)\
                            + "_"\
                            + name_BR\
                            + str(BR)\
                            + "."\
                            + ws.cell(row=i+counter*number_row,column=3).value
                    counter += 1
    ws.insert_cols(2)
    wb.save(path_to_new_template)